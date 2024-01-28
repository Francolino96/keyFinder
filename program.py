import csv
import os
import pandas as pd
import openpyxl
from docx import Document
import PyPDF2


# Funzione che legge dal file chiavi_da_cercare.csv le chiavi che andranno cercate
def read_names_from_csv(file_path):
    names = []
    try:
        with open(file_path, 'r') as csvfile:
            csvreader = csv.reader(csvfile)
            names = [row[0] for row in csvreader]

    except FileNotFoundError:
        print(f"Errore: File non trovato in {file_path}")
    except Exception as e:
        print(f"Errore: {e}")
    return names


# Funzione che cerca tutte le chiavi
def search_key_in_folder_recursive(response_path, directory, key):
    if not os.path.exists(directory):
        print(f"Errore: la directory '{directory}' non esiste.")
        return

    for filename in os.listdir(directory):
        path = os.path.join(directory, filename)
        if os.path.isdir(path):
            search_key_in_folder_recursive(response_path, path, key)
        elif filename.endswith(".csv"):
            search_key_in_csv(response_path, path, key)
        elif filename.endswith(".pdf"):
            search_key_in_pdf(response_path, path, key)
        elif filename.endswith(".txt"):
            search_key_in_txt(response_path, path, key)
        elif filename.endswith(".xls"):
            search_key_in_xls(response_path, path, key)
        elif filename.endswith(".xlsx"):
            search_key_in_xlsx(response_path, path, key)
        elif filename.endswith(".docx"):
            search_key_in_docx(response_path, path, key)
        elif filename.endswith(".htm") | filename.endswith(".html"):
            search_key_in_htm(response_path, path, key)


# Funzione di ricerca chiave per files di tipo .csv
def search_key_in_csv(response_path, csv_filepath, key):
    try:
        with open(csv_filepath, 'r', encoding='utf-8') as csvfile:
            csvreader = csv.reader(csvfile)
            for row_number, row in enumerate(csvreader, start=1):
                if any(key in cell for cell in row):
                    with open(response_path, 'a') as file:
                        file.write(f"  '{csv_filepath}'\n")
                    return

    except FileNotFoundError:
        print(f"Errore: File non trovato in {csv_filepath}")
    except Exception as e:
        print(f"Errore: {e}")


# Funzione di ricerca chiave per files di tipo .pdf
def search_key_in_pdf(response_path, pdf_filepath, key):
    try:
        with open(pdf_filepath, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            for page_number in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_number]
                page_text = page.extract_text()
                if key in page_text:
                    with open(response_path, 'a') as file:
                        file.write(f"  '{pdf_filepath}'\n")
                    return

    except FileNotFoundError:
        print(f"Errore: File non trovato in {pdf_filepath}")
    except Exception as e:
        print(f"Errore: {e}")


# Funzione di ricerca chiave per files di tipo .txt
def search_key_in_txt(response_path, txt_filepath, key):
    try:
        with open(txt_filepath, 'r', encoding='utf-8') as txt_file:
            txt_content = txt_file.read()
            if key in txt_content:
                with open(response_path, 'a') as file:
                    file.write(f"  '{txt_filepath}'\n")
                return

    except FileNotFoundError:
        print(f"Errore: File non trovato in {txt_filepath}")
    except Exception as e:
        print(f"Errore: {e}")


# Funzione di ricerca chiave per files di tipo .xls
def search_key_in_xls(response_path, xls_filepath, key):
    try:
        df = pd.read_excel(xls_filepath, header=None)
        for index, row in df.iterrows():
            for column, value in row.items():
                if pd.notna(value) and key in str(value):
                    with open(response_path, 'a') as file:
                        file.write(f"  '{xls_filepath}'\n")
                    return

    except FileNotFoundError:
        print(f"Errore: File non trovato in {xls_filepath}")
    except Exception as e:
        print(f"Errore: {e}")


# Funzione di ricerca chiave per files di tipo .xlsx
def search_key_in_xlsx(response_path, xlsx_filepath, key):
    try:
        workbook = openpyxl.load_workbook(xlsx_filepath)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if key in str(cell.value):
                       with open(response_path, 'a') as file:
                            file.write(f"  '{xlsx_filepath}'\n")
                       return

    except FileNotFoundError:
        print(f"Errore: File non trovato in {xlsx_filepath}")
    except Exception as e:
        print(f"Errore: {e}")


# Funzione di ricerca chiave per files di tipo .docx
def search_key_in_docx(response_path, docx_filepath, key):
    try:
        docx = Document(docx_filepath)
        for paragraph_number, paragraph in enumerate(docx.paragraphs, start=1):
            if key in paragraph.text:
                with open(response_path, 'a') as file:
                    file.write(f"  '{docx_filepath}'\n")
                return

    except FileNotFoundError:
        print(f"Errore: File non trovato in {docx_filepath}")
    except Exception as e:
        print(f"Errore: {e}")


# Funzione di ricerca chiave per files di tipo .htm o .html
def search_key_in_htm(response_path, htm_filepath, key):
    try:
        with open(htm_filepath, 'r', encoding='utf-8') as htm_file:
            htm_content = htm_file.read()
            if key in htm_content:
                with open(response_path, 'a') as file:
                    file.write(f"  '{htm_filepath}'\n")
                return

    except FileNotFoundError:
        print(f"Errore: File non trovato in {htm_filepath}")
    except Exception as e:
        print(f"Errore: {e}")


if __name__ == '__main__':
    project_path = '.\\'
    all_files = os.listdir(project_path)
    csv_files = [file for file in all_files if file.endswith('.csv')]
    if not csv_files:
        print(f"Errore: non e' presente al lista di chiavi nella cartella corrente '{project_path}'. Accertarsi che il file abbia estensione .csv")
    first_csv_file = csv_files[0]
    csv_file_path = os.path.join(project_path, first_csv_file)
    names_vector = read_names_from_csv(csv_file_path)

    print("\n\nRICERCA CHIAVI IN CORSO...")

    response_file_path = 'lista_occorrenze.txt'

    with open(response_file_path, 'w') as file:
        file.write("\n\n------------------------ FILE DI RISPOSTA --------------------------\n")
        file.write("Estensioni supportate: .xls, .xlsx, .txt, .csv, .docx, .pdf, .html, .htm\n")

    base_path = '..\\'
    all_folders = [folder for folder in os.listdir(base_path) if
                   os.path.isdir(os.path.join(base_path, folder))]
    folders_to_process = [folder for folder in all_folders if folder != 'pythonProject']
    for name in names_vector:
        print("CHIAVE CERCATA: '" + name + "'")
        with open(response_file_path, 'a') as file:
            file.write("\n\n\nCHIAVE CERCATA: '" + name + "'\n\n")
            file.write("FILE IN CUI VI E' ALMENO UN'OCCORRENZA:\n")
        for folder in folders_to_process:
            folder_path = os.path.join(base_path, folder)
            search_key_in_folder_recursive(response_file_path, folder_path, name)
